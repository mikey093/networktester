import sys
import os
import json
import subprocess
import csv
from time import strftime
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from pythonping import ping
import logging


with open('config/testConfig.json', 'r') as file:
    jsonData = json.load(file)


usage = '''
Usage: 
stress.py [Options]
Options:
-f [File Name] 
-b [TCP bandwidth(kb/s)] 
-u [UDP bandwidth(kb/s)] 
-h [Iperf Server IP Address] 
-t [time of test (sec)] 
-p [Port for Iperf]
-n [No. of Tests]
-c [Comment of location]
'''

if("-f" in sys.argv):
    projectFile = 'results/{}'.format(
        sys.argv[sys.argv.index("-f") + 1])
    logFile = 'log/{}.log'.format(sys.argv[sys.argv.index("-f") + 1])
else:
    print(usage)
    print("Specify output file.")
    exit()
if("-b" in sys.argv):
    bandTCP = sys.argv[sys.argv.index("-b") + 1]
else:
    bandTCP = 10
if("-u" in sys.argv):
    bandUDP = sys.argv[sys.argv.index("-u") + 1]
else:
    bandUDP = 1
if("-h" in sys.argv):
    iperfHost = sys.argv[sys.argv.index("-h") + 1]
else:
    print(usage)
    exit()
if("-p" in sys.argv):
    iperfPort = sys.argv[sys.argv.index("-p") + 1]
else:
    iperfPort = 8071
    print("Setting iperf port to 8071")
if("-t" in sys.argv):
    iperfTime = sys.argv[sys.argv.index("-t") + 1]
else:
    print(usage)
    exit()
if("-n" in sys.argv):
    numTests = sys.argv[sys.argv.index("-n") + 1]
else:
    print(usage)
    exit()
if("-c" in sys.argv):
    locationComment = sys.argv[sys.argv.index("-c") + 1]
else:
    locationComment = ''

# Setup logging.  
logging.basicConfig(filename=logFile, level=logging.DEBUG)
logging.getLogger().addHandler(logging.StreamHandler())


# Provide test settings info for user.
testSettings = '''
Settings for test:
File Name: {},
TCP bandwidth(kb/s): {},
UDP bandwidth(kb/s): {},
Iperf Server IP Address: {},
Iperf Server port: {},
time of test (sec): {},
No. of Tests: {}
'''.format(projectFile, bandTCP, bandUDP, iperfHost, iperfPort, iperfTime, numTests)

# print(testSettings)
logging.info('Test settings at {}: {}'.format(time.strftime("%d/%m/%Y %X"), testSettings))

def stressTest(jsonData, testNumber=1):
    iperfSaveFileTCP = 'results/TCP-{}.json'.format(testNumber)
    iperfSaveFileUDP = 'results/UDP-{}.json'.format(testNumber)

    iperfCommandTCP = ('iperf3.exe -c {} -J -b {} -t {} -p {} > {}'.format(
        iperfHost, int(bandTCP)*1000, iperfTime, iperfPort, iperfSaveFileTCP))

    unlimited_iperfCommandUDP = ('iperf3.exe -c {} -u -J -b {} -t {} -p {} > {}'.format(
        iperfHost, int(bandUDP)*1000, iperfTime, iperfPort, iperfSaveFileUDP))

    try:
        subprocess.call(str(iperfCommandTCP), shell=True,
                        timeout=int(iperfTime) + 2)
    except subprocess.TimeoutExpired as errorMessage:
        logging.info(errorMessage)
        resultsDict = 0
        return resultsDict
    try:
        subprocess.call(str(unlimited_iperfCommandUDP),
                        shell=True, timeout=int(iperfTime) + 2)
    except subprocess.TimeoutExpired as errorMessage:
        logging.info(errorMessage)
        resultsDict = 0
        return resultsDict

    with open(iperfSaveFileTCP, 'r') as raw:
        rawTCP = json.load(raw)
    # I am assuming that if the test doesn't run correct, we'll always get a KeyError here.
    # Will have to add extra tests so we can return better error messages.
    try:
        results = rawTCP['end']['streams'][0]['sender']
    except KeyError:
        results = {}
        logging.info('Error - Did not successfully run TCP test')
        resultsDict = 0

    with open(iperfSaveFileUDP, 'r') as raw:
        rawUDP = json.load(raw)
    try:
        resultsUDP = rawUDP['end']['streams'][0]['udp']
    except KeyError:
        resultsUDP = {}
        logging.info('Error - Did not successfully run UDP test')
        resultsDict = 0

    try:
        resultsCPU = rawTCP['end']['cpu_utilization_percent']
        generalResults = rawTCP['start']['connected'][0]
    except KeyError:
        logging.info('Test did not run - Cannot get CPU results')
        resultsCPU = 'NA'
        generalResults = 'NA'

    timeNow = time.strftime("%d/%m/%Y %X")
    try:
        bandwidth = round(results['bits_per_second'] * (1e-6), 2)
        timeOfTest = round(results['seconds'], 1) + \
            round(resultsUDP['seconds'], 1)
        totalBytes = results['bytes']
        hostCPU = round(resultsCPU['host_total'], 1)
        remoteCPU = round(resultsCPU['remote_total'], 1)
        localIP = generalResults['local_host']
        localPort = generalResults['local_port']
        logging.info("TCP Bandwidth = {} Mb/s".format(bandwidth))
    except KeyError:
        logging.info('Test did not run correctly.  Setting NA values.')
        bandwidth = 'NA'
        timeOfTest = 'NA'
        totalBytes = 'NA'
        hostCPU = 'NA'
        remoteCPU = 'NA'
        localIP = 'NA'
        localPort = 'NA'

    try:
        udpResults = {
            'UDP Bandwidth (Mb/s)': round(resultsUDP['bits_per_second'] * (1e-6), 2),
            'Jitter (ms)': resultsUDP['jitter_ms'],
            'Total UDP Packets': resultsUDP['packets'],
            'Lost Packets': resultsUDP['lost_packets'],
            'Lost Packets (%)': resultsUDP['lost_percent'],
            'Packets Out of Order': resultsUDP['out_of_order']
        }
    except KeyError:
        udpResults = {
            'UDP Bandwidth (Mb/s)': 'NA',
            'Jitter (ms)': 'NA',
            'Total UDP Packets': 'NA',
            'Lost Packets': 'NA',
            'Lost Packets (%)': 'NA',
            'Packets Out of Order': 'NA'
        }

    # RTT doesn't work on my Windows PC...
    try:
        RTT_ms = results['max_rtt'] * (1e-3)
    except:
        logging.info('RTT Error.')
        RTT_ms = 'NA'
    try:
        pingResults = ping(iperfHost, size=40, count=10)
        latencyAvg = pingResults.rtt_avg_ms
    except:
        latencyAvg = 'NA'

    resultsDict = {
        'Local IP Address': localIP,
        'Server Address': iperfHost,
        'Local Port': localPort,
        'Server Port': iperfPort,
        'Time': timeNow,
        'Bandwidth (Mb/s)': bandwidth,
        'RTT (ms)': RTT_ms,
        'UDP Bandwidth (Mb/s)': udpResults['UDP Bandwidth (Mb/s)'],
        'Jitter (ms)': udpResults['Jitter (ms)'],
        'Total UDP Packets': udpResults['Total UDP Packets'],
        'Lost Packets': udpResults['Lost Packets'],
        'Lost Packets (%)': udpResults['Lost Packets (%)'],
        'Packets Out of Order': udpResults['Packets Out of Order'],
        'Test Duration (sec)': timeOfTest,
        'Total Bytes': totalBytes,
        'Host CPU Utilisation (%)': hostCPU,
        'Remote CPU Utilisation (%)': remoteCPU,
        'Latency (ms)': latencyAvg,
        'Comment': locationComment
    }
    # print(resultsDict)
    return resultsDict


def toCSV(resultsDict):
    # Grab dictionary and put into CSV
    # print(resultsDict)
    headerOfCSV = list(resultsDict.keys())
    csvFileOut = '{}.csv'.format(projectFile)
    # Look for file - If it doesn't exist, write header.
    # If file not found, append results.
    # This is still not opening in Excel properly..

    if os.path.isfile(csvFileOut) == False:
        logging.info('No Header.  Adding header to {}'.format(csvFileOut))
        with open(csvFileOut, 'w', newline='') as outFile:
            writer = csv.DictWriter(
                outFile, fieldnames=headerOfCSV, dialect='excel')
            writer.writeheader()
            writer.writerow(resultsDict)
    else:
        with open(csvFileOut, 'a', newline='') as outFile:
            writer = csv.DictWriter(
                outFile, fieldnames=headerOfCSV, dialect='excel')
            writer.writerow(resultsDict)


def conditions(theValue, threshold, passMark, inequality=0):
    # Iterate through each value going into sheet.
    # FF0000 = Red.
    # Must be a better way to do this..
    
    if inequality == 1:
        if theValue > threshold:
            theFont = Font(color='00FF00')
        elif theValue < threshold:
            theFont = Font(color='FF0000')
            passMark = 0
    else:
        if theValue < threshold:
            theFont = Font(color='00FF00')
        elif theValue > threshold:
            theFont = Font(color='FF0000')
            passMark = 0

    return theFont, passMark


def toXL(filePrefix):
    # Puts CSV into Excel with formatted values depending on their value.
    # Downloaded from results page.
    thresholdsFile = 'config/thresholds.json'
    csvFile = '{}.csv'.format(filePrefix)
    xlsxFile = '{}.xlsx'.format(filePrefix)
    projectFile = 'config/projectInfo.json'

    try:
        wb = load_workbook(filename=('config/template.xlsx'))
    except PermissionError:
        logging.info('Close Worksheet.')
    except:
        logging.info('Something is wrong with loading workbook!')
    # Check for CSV file presence.
    ws = wb.active
    with open(thresholdsFile, 'r') as openFile:
        thresholdsData = json.load(openFile)
    with open(projectFile, 'r') as projectInfo:
        proj = json.load(projectInfo)
    with open(csvFile, 'r') as csv_input:
        results = csv.DictReader(csv_input)

        # Start at 11 (where the spreadsheet starts inputs.)
        i = 11
        # All the thresholds.
        ws['C{}'.format(i-1)] = '> {}'.format(thresholdsData['TCPbandwidth'])
        ws['D{}'.format(i-1)] = '> {}'.format(thresholdsData['UDPbandwidth'])
        #ws['E{}'.format(i-1)] = '< {}'.format(thresholdsData['RTT'])
        ws['E{}'.format(i-1)] = '< {}'.format(thresholdsData['jitter'])
        # ws['G{}'.format(i-1)] = '< {}'.format(thresholdsData['RSS'])
        # ws['H{}'.format(i-1)] = '< {}'.format(thresholdsData['linkSpeed'])
        ws['F{}'.format(i-1)] = '< {}'.format(thresholdsData['latency'])
        ws['G{}'.format(i-1)] = '< {}'.format(thresholdsData['packetLoss'])
        
        # Additional testing parameters.
        ws['O2'] = '{} seconds for each test.'.format(iperfTime)
        ws['O3'] = '{} Mb/s'.format(int(bandTCP)/1000)
        ws['O4'] = '{} Mb/s'.format(int(bandUDP)/1000)
        ws['O5'] = '{}'.format(iperfPort)

        # Project Info
        ws['B3'] = proj['Serial Number']
        ws['B4'] = proj['MAC Address']
        ws['B6'] = proj['Location Name']
        ws['B7'] = proj['Engineer Name']
        

        # ws['B5'] = configs['host']

        for row in results:
            passMark = 1
            ws['A{}'.format(i)] = i - 10
            ws['B{}'.format(i)] = row['Time']

            # I think I've fucked this up.  It's so complex going through the conditions statement..
            try:
                ws['C{}'.format(i)] = float(row['Bandwidth (Mb/s)'])
                ws['C{}'.format(i)].font, passMark = conditions(
                    float(row['Bandwidth (Mb/s)']), thresholdsData['TCPbandwidth'], passMark, 1)
            except ValueError:
                ws['C{}'.format(i)] = row['Bandwidth (Mb/s)']
                ws['C{}'.format(i)].font = Font(color='FF0000')
                passMark = 0
                # ws['C{}'.format(i)].font = 'FF000000'

            try:
                ws['D{}'.format(i)] = float(row['UDP Bandwidth (Mb/s)'])
                ws['D{}'.format(i)].font, passMark = conditions(
                    float(row['UDP Bandwidth (Mb/s)']), thresholdsData['UDPbandwidth'], passMark, 1)
            except ValueError:
                ws['D{}'.format(i)] = row['UDP Bandwidth (Mb/s)']
                ws['D{}'.format(i)].font = Font(color='FF0000')
                passMark = 0
            # # ws['K{}'.format(i)] = row['Description of Location']

            try:
                ws['E{}'.format(i)] = float(row['Jitter (ms)'])
                ws['E{}'.format(i)].font, passMark = conditions(
                    float(row['Jitter (ms)']), thresholdsData['jitter'], passMark)
            except ValueError:
                ws['E{}'.format(i)] = row['Jitter (ms)']
                ws['E{}'.format(i)].font = Font(color='FF0000')
                passMark = 0            
            
            try:
                ws['G{}'.format(i)] = float(row['Lost Packets (%)'])
                ws['G{}'.format(i)].font, passMark = conditions(
                    float(row['Lost Packets (%)']), 1, passMark)
            except ValueError:
                ws['G{}'.format(i)] = row['Lost Packets (%)']
                ws['G{}'.format(i)].font = Font(color='FF0000')
                passMark = 0            


            # Did this condition a bit different just to try.  If not a float it will go through conditions.  Probably better.
            if row['Latency (ms)'] == 'NA':
                ws['F{}'.format(i)] = (row['Latency (ms)'])
            else:
                ws['F{}'.format(i)] = float(row['Latency (ms)'])
                ws['F{}'.format(i)].font, passMark = conditions(
                    float(row['Latency (ms)']), thresholdsData['latency'], passMark)


            ws['H{}'.format(i)] = row['Comment']

            if passMark == 1:
                ws['I{}'.format(i)] = 'X'
            else:
                ws['J{}'.format(i)] = 'X'

            i += 1
        # except Exception as e:
        #     print(e)
        # except ValueError as valueerrorXL:
        #     print('bad')
        #     logging.info(valueerrorXL)
    logging.info('Writing spreadsheet.')
    try:
        wb.save(xlsxFile)
    except PermissionError:
        logging.error('Could not write to file.  Excel sheet is probably open.  Close and try again.')

x = int(numTests)
while x >= 1:
    print('---------------------------')
    print('Running the following test:')
    print('{} kb/s TCP and {} kb/s UDP to {} for {} seconds'.format(
        bandTCP, bandUDP, iperfHost, iperfTime))
    print('---------------------------')
    resultsDict = stressTest(jsonData)
    toCSV(resultsDict)
    print('---------------------------')
    time.sleep(2)
    x = x - 1

toXL(projectFile)
# Spreadsheet write.

