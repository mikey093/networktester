import sys
import json
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font
import logging

version = open('VERSION.txt', 'r').read()
if("-v" in sys.argv):
    print('Version: {}'.format(version))
    sys.exit(0)

if("-f" in sys.argv):
    projectFile = 'results/{}'.format(
        sys.argv[sys.argv.index("-f") + 1])
    logFile = 'log/{}.log'.format(sys.argv[sys.argv.index("-f") + 1])
else:
    print(usage)
    print("Specify file.")
    sys.exit()

# Setup logging.
logging.basicConfig(filename=logFile, level=logging.DEBUG)
logging.getLogger().addHandler(logging.StreamHandler())


def conditions(theValue, threshold, passMark, inequality=0):
    # Iterate through each value going into sheet.
    # FF0000 = Red.
    # Must be a better way to do this..

    if inequality == 1:
        if theValue > threshold:
            theFont = Font(color='00FF00')
        elif theValue < threshold:
            theFont = Font(color='FF0000')
            passMark = 0
    else:
        if theValue < threshold:
            theFont = Font(color='00FF00')
        elif theValue > threshold:
            theFont = Font(color='FF0000')
            passMark = 0

    return theFont, passMark


def toXL(filePrefix):
    # Puts CSV into Excel with formatted values depending on their value.
    # Downloaded from results page.
    thresholdsFile = 'config/thresholds.json'
    csvFile = '{}.csv'.format(filePrefix)
    xlsxFile = '{}.xlsx'.format(filePrefix)
    projectInfo = 'config/projectInfo.json'

    try:
        wb = load_workbook(filename=('config/template.xlsx'))
    except PermissionError:
        logging.info('Close Worksheet.')
    except:
        logging.info('Something is wrong with loading workbook!')
    # Check for CSV file presence.
    ws = wb.active
    with open(thresholdsFile, 'r') as openFile:
        thresholdsData = json.load(openFile)
    with open(projectInfo, 'r') as projectInfo:
        proj = json.load(projectInfo)
    with open(csvFile, 'r') as csv_input:
        results = csv.DictReader(csv_input)

        # Start at 11 (where the spreadsheet starts inputs.)
        i = 11
        # All the thresholds.
        ws['C{}'.format(i-1)] = '> {}'.format(thresholdsData['TCPbandwidth'])
        ws['D{}'.format(i-1)] = '> {}'.format(thresholdsData['UDPbandwidth'])
        #ws['E{}'.format(i-1)] = '< {}'.format(thresholdsData['RTT'])
        ws['E{}'.format(i-1)] = '< {}'.format(thresholdsData['jitter'])
        # ws['G{}'.format(i-1)] = '< {}'.format(thresholdsData['RSS'])
        # ws['H{}'.format(i-1)] = '< {}'.format(thresholdsData['linkSpeed'])
        ws['F{}'.format(i-1)] = '< {}'.format(thresholdsData['latency'])
        ws['G{}'.format(i-1)] = '< {}'.format(thresholdsData['packetLoss'])

        # Project Info
        ws['B7'] = proj['Engineer Name']
        # ws['B5'] = configs['host']

        for row in results:
            passMark = 1
            ws['A{}'.format(i)] = i - 10
            ws['B{}'.format(i)] = row['Time']

            # Additional testing parameters.
            ws['B2'] = '{} seconds for each test.'.format(
                row['Test Seconds (s)'])
            ws['B3'] = '{} Mb/s'.format(float(row['Test TCP Bandwidth (Mb/s)']))
            ws['B4'] = '{} Mb/s'.format(float(row['Test UDP Bandwidth (Mb/s)']))
            ws['B5'] = '{}'.format(row['Test iperf port'])

            # I think I've fucked this up.  It's so complex going through the conditions statement..
            try:
                ws['C{}'.format(i)] = float(row['Bandwidth (Mb/s)'])
                ws['C{}'.format(i)].font, passMark = conditions(
                    float(row['Bandwidth (Mb/s)']), thresholdsData['TCPbandwidth'], passMark, 1)
            except ValueError:
                ws['C{}'.format(i)] = row['Bandwidth (Mb/s)']
                ws['C{}'.format(i)].font = Font(color='FF0000')
                passMark = 0
                # ws['C{}'.format(i)].font = 'FF000000'

            try:
                ws['D{}'.format(i)] = float(row['UDP Bandwidth (Mb/s)'])
                ws['D{}'.format(i)].font, passMark = conditions(
                    float(row['UDP Bandwidth (Mb/s)']), thresholdsData['UDPbandwidth'], passMark, 1)
            except ValueError:
                ws['D{}'.format(i)] = row['UDP Bandwidth (Mb/s)']
                ws['D{}'.format(i)].font = Font(color='FF0000')
                passMark = 0
            # # ws['K{}'.format(i)] = row['Description of Location']

            try:
                ws['E{}'.format(i)] = float(row['Jitter (ms)'])
                ws['E{}'.format(i)].font, passMark = conditions(
                    float(row['Jitter (ms)']), thresholdsData['jitter'], passMark)
            except ValueError:
                ws['E{}'.format(i)] = row['Jitter (ms)']
                ws['E{}'.format(i)].font = Font(color='FF0000')
                passMark = 0

            try:
                ws['G{}'.format(i)] = float(row['Lost Packets (%)'])
                ws['G{}'.format(i)].font, passMark = conditions(
                    float(row['Lost Packets (%)']), 1, passMark)
            except ValueError:
                ws['G{}'.format(i)] = row['Lost Packets (%)']
                ws['G{}'.format(i)].font = Font(color='FF0000')
                passMark = 0

            # Did this condition a bit different just to try.  If not a float it will go through conditions.  Probably better.
            if row['Latency (ms)'] == 'NA':
                ws['F{}'.format(i)] = (row['Latency (ms)'])
            else:
                ws['F{}'.format(i)] = float(row['Latency (ms)'])
                ws['F{}'.format(i)].font, passMark = conditions(
                    float(row['Latency (ms)']), thresholdsData['latency'], passMark)

            ws['H{}'.format(i)] = row['Comment']

            if passMark == 1:
                ws['I{}'.format(i)] = 'X'
            else:
                ws['J{}'.format(i)] = 'X'

            i += 1

    logging.info('Writing spreadsheet.')
    try:
        wb.save(xlsxFile)
    except PermissionError:
        logging.error(
            'Could not write to file.  Excel sheet is probably open.  Close and try again.')


toXL(projectFile)
# Spreadsheet write.
